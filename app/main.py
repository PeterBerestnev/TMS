from datetime import date, datetime, timedelta
from typing import Optional

from fastapi import Depends, FastAPI, Form, HTTPException, Request
from fastapi.responses import HTMLResponse, RedirectResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from sqlalchemy import func, select
from sqlalchemy.orm import Session

from .database import Base, engine, get_db
from .models import Category, TimeEntry

app = FastAPI(title="Time Management Tracker")

Base.metadata.create_all(bind=engine)

app.mount("/static", StaticFiles(directory="app/static"), name="static")
templates = Jinja2Templates(directory="app/templates")


def get_today() -> date:
    return date.today()


@app.get("/", response_class=HTMLResponse)
async def index(
    request: Request,
    db: Session = Depends(get_db),
    day: Optional[date] = None,
):
    if day is None:
        day = get_today()

    categories = db.execute(select(Category).order_by(Category.name)).scalars().all()
    entries = (
        db.execute(
            select(TimeEntry)
            .where(TimeEntry.date == day)
            .order_by(TimeEntry.start_time, TimeEntry.id)
        )
        .scalars()
        .all()
    )

    return templates.TemplateResponse(
        "index.html",
        {
            "request": request,
            "day": day,
            "entries": entries,
            "categories": categories,
        },
    )


@app.post("/entries/add")
async def add_entry(
    date_value: date = Form(...),
    category_id: int = Form(...),
    duration_hours: float = Form(...),
    comment: str = Form(""),
    start_time_str: str = Form(""),
    end_time_str: str = Form(""),
    db: Session = Depends(get_db),
):
    category = db.get(Category, category_id)
    if not category:
        raise HTTPException(status_code=404, detail="Category not found")

    start_dt: Optional[datetime] = None
    end_dt: Optional[datetime] = None
    if start_time_str:
        start_dt = datetime.combine(date_value, datetime.strptime(start_time_str, "%H:%M").time())
    if end_time_str:
        end_dt = datetime.combine(date_value, datetime.strptime(end_time_str, "%H:%M").time())

    entry = TimeEntry(
        date=date_value,
        category_id=category_id,
        duration_hours=duration_hours,
        comment=comment or None,
        start_time=start_dt,
        end_time=end_dt,
    )
    db.add(entry)
    db.commit()

    return RedirectResponse(url=f"/?day={date_value.isoformat()}", status_code=303)


@app.post("/entries/{entry_id}/delete")
async def delete_entry(entry_id: int, db: Session = Depends(get_db)):
    entry = db.get(TimeEntry, entry_id)
    if not entry:
        raise HTTPException(status_code=404, detail="Entry not found")
    date_value = entry.date
    db.delete(entry)
    db.commit()
    return RedirectResponse(url=f"/?day={date_value.isoformat()}", status_code=303)


@app.get("/categories", response_class=HTMLResponse)
async def categories_page(request: Request, db: Session = Depends(get_db)):
    categories = db.execute(select(Category).order_by(Category.name)).scalars().all()
    return templates.TemplateResponse(
        "categories.html",
        {"request": request, "categories": categories},
    )


@app.post("/categories/add")
async def add_category(
    name: str = Form(...),
    color: str = Form("#1976d2"),
    description: str = Form(""),
    db: Session = Depends(get_db),
):
    exists = db.execute(select(Category).where(Category.name == name)).scalar_one_or_none()
    if exists:
        raise HTTPException(status_code=400, detail="Category with this name already exists")

    category = Category(name=name, color=color or None, description=description or None)
    db.add(category)
    db.commit()
    return RedirectResponse(url="/categories", status_code=303)


@app.post("/categories/{category_id}/delete")
async def delete_category(category_id: int, db: Session = Depends(get_db)):
    category = db.get(Category, category_id)
    if not category:
        raise HTTPException(status_code=404, detail="Category not found")
    db.delete(category)
    db.commit()
    return RedirectResponse(url="/categories", status_code=303)


@app.get("/calendar", response_class=HTMLResponse)
async def calendar_view(
    request: Request,
    db: Session = Depends(get_db),
    month: Optional[int] = None,
    year: Optional[int] = None,
):
    today = get_today()
    if month is None:
        month = today.month
    if year is None:
        year = today.year

    start_date = date(year, month, 1)
    if month == 12:
        next_month = date(year + 1, 1, 1)
    else:
        next_month = date(year, month + 1, 1)
    end_date = next_month - timedelta(days=1)

    days = []
    current = start_date
    while current <= end_date:
        days.append(current)
        current += timedelta(days=1)

    totals = (
        db.execute(
            select(TimeEntry.date, func.sum(TimeEntry.duration_hours))
            .where(TimeEntry.date >= start_date, TimeEntry.date <= end_date)
            .group_by(TimeEntry.date)
            .order_by(TimeEntry.date)
        )
        .all()
    )
    totals_map = {d: h for d, h in totals}

    return templates.TemplateResponse(
        "calendar.html",
        {
            "request": request,
            "month": month,
            "year": year,
            "days": days,
            "totals_map": totals_map,
        },
    )


@app.get("/stats", response_class=HTMLResponse)
async def stats_view(
    request: Request,
    db: Session = Depends(get_db),
    start: Optional[date] = None,
    end: Optional[date] = None,
):
    today = get_today()
    if end is None:
        end = today
    if start is None:
        start = end - timedelta(days=6)

    per_category = (
        db.execute(
            select(Category.name, func.sum(TimeEntry.duration_hours))
            .join(TimeEntry, TimeEntry.category_id == Category.id)
            .where(TimeEntry.date >= start, TimeEntry.date <= end)
            .group_by(Category.name)
            .order_by(Category.name)
        )
        .all()
    )

    cat_labels = [name for name, _ in per_category]
    cat_values = [float(hours or 0) for _, hours in per_category]

    per_day = (
        db.execute(
            select(TimeEntry.date, func.sum(TimeEntry.duration_hours))
            .where(TimeEntry.date >= start, TimeEntry.date <= end)
            .group_by(TimeEntry.date)
            .order_by(TimeEntry.date)
        )
        .all()
    )

    day_labels = [d.isoformat() for d, _ in per_day]
    day_values = [float(hours or 0) for _, hours in per_day]

    return templates.TemplateResponse(
        "stats.html",
        {
            "request": request,
            "start": start,
            "end": end,
            "per_category": per_category,
            "per_day": per_day,
            "cat_labels": cat_labels,
            "cat_values": cat_values,
            "day_labels": day_labels,
            "day_values": day_values,
        },
    )


@app.get("/export/excel")
async def export_excel(
    db: Session = Depends(get_db),
    start: Optional[date] = None,
    end: Optional[date] = None,
):
    from io import BytesIO

    import openpyxl
    from openpyxl.utils import get_column_letter

    today = get_today()
    if end is None:
        end = today
    if start is None:
        start = end - timedelta(days=30)

    entries = (
        db.execute(
            select(TimeEntry, Category)
            .join(Category, TimeEntry.category_id == Category.id)
            .where(TimeEntry.date >= start, TimeEntry.date <= end)
            .order_by(TimeEntry.date, TimeEntry.id)
        )
        .all()
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Time entries"

    headers = [
        "Дата",
        "Категория",
        "Часы",
        "Комментарий",
        "Время начала",
        "Время окончания",
    ]
    ws.append(headers)

    for time_entry, category in entries:
        ws.append(
            [
                time_entry.date.isoformat(),
                category.name,
                time_entry.duration_hours,
                time_entry.comment or "",
                time_entry.start_time.time().strftime("%H:%M") if time_entry.start_time else "",
                time_entry.end_time.time().strftime("%H:%M") if time_entry.end_time else "",
            ]
        )

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].auto_size = True

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)

    filename = f"time_entries_{start.isoformat()}_{end.isoformat()}.xlsx"
    headers_resp = {
        "Content-Disposition": f'attachment; filename="{filename}"'
    }

    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers_resp,
    )

